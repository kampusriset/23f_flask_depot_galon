<<<<<<< HEAD
from flask import Flask, render_template, request, redirect, url_for, session, flash
=======
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file, make_response
>>>>>>> 4fd9892 (final code)
from datetime import datetime, timedelta
import os
import json
import pymysql
<<<<<<< HEAD

# Import dari database folder
=======
import random
from io import BytesIO

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("Warning: openpyxl tidak terinstall. Fitur Excel akan dinonaktifkan.")

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    print("Warning: reportlab tidak terinstall. Fitur PDF akan dinonaktifkan.")

>>>>>>> 4fd9892 (final code)
from database.models import db, User, Pemesanan, Galon, PembelianGalon, Pembayaran, Laporan, Pengaturan

app = Flask(__name__)
app.secret_key = 'secretkey123'

<<<<<<< HEAD
# ⭐⭐⭐ UBAH INI ⭐⭐⭐
# DARI: 'sqlite:///depot_air.db'
# MENJADI:
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@localhost/depot_air'
# Penjelasan: root = username, kosong = password, localhost = server, depot_air = nama database
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# Data default untuk galon
=======
app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql+pymysql://root:@localhost/depot_air'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

>>>>>>> 4fd9892 (final code)
DEFAULT_GALON = [
    {"jenis": "Air Galon Biasa", "harga": 6000, "stok": 100},
    {"jenis": "Air Galon RO", "harga": 8000, "stok": 100},
    {"jenis": "Air Galon Mineral", "harga": 7000, "stok": 100}
]

@app.route('/')
def home():
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        nama_lengkap = request.form.get('nama_lengkap', '')
        email = request.form.get('email', '')
        telepon = request.form.get('telepon', '')
        
<<<<<<< HEAD
        # Cek jika username sudah ada
=======
>>>>>>> 4fd9892 (final code)
        existing_user = User.query.filter_by(username=username).first()
        if existing_user:
            flash('Username sudah digunakan', 'error')
            return redirect(url_for('register'))
        
<<<<<<< HEAD
        # Buat user baru
=======
>>>>>>> 4fd9892 (final code)
        new_user = User(
            username=username, 
            nama_lengkap=nama_lengkap, 
            email=email,
            telepon=telepon
        )
        new_user.set_password(password)
        
        try:
            db.session.add(new_user)
            db.session.commit()
            flash('Registrasi berhasil! Silakan login.', 'success')
            return redirect(url_for('login'))
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan saat registrasi: {str(e)}', 'error')
            return redirect(url_for('register'))
    
    return render_template('register.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
<<<<<<< HEAD
        
        user = User.query.filter_by(username=username).first()
        
=======
        login_type = request.form.get('login_type', 'customer')
        
        user = User.query.filter_by(username=username).first()
        
        if not user:
            flash('Username atau password salah', 'error')
            return redirect(url_for('login'))
        
        if login_type == 'admin' and user.role != 'admin':
            flash('Akun ini bukan admin. Silakan login sebagai pelanggan.', 'error')
            return redirect(url_for('login'))
        
>>>>>>> 4fd9892 (final code)
        if user and user.check_password(password) and user.is_active:
            session['user_id'] = user.id
            session['username'] = user.username
            session['nama_lengkap'] = user.nama_lengkap or user.username
            session['role'] = user.role
            
            flash('Login berhasil!', 'success')
            
<<<<<<< HEAD
            # Redirect berdasarkan role
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                return redirect(url_for('dashboard'))
=======
            if user.role == 'admin':
                return redirect(url_for('admin_dashboard'))
            else:
                # Set flag untuk menampilkan welcome page
                session['show_welcome'] = True
                return redirect(url_for('welcome'))
>>>>>>> 4fd9892 (final code)
        else:
            flash('Username atau password salah', 'error')
            return redirect(url_for('login'))
    
    return render_template('login.html')

<<<<<<< HEAD
@app.route('/dashboard')
def dashboard():
=======
@app.route('/welcome')
def welcome():
>>>>>>> 4fd9892 (final code)
    if 'user_id' not in session:
        flash('Silakan login terlebih dahulu', 'error')
        return redirect(url_for('login'))
    
<<<<<<< HEAD
    user_id = session['user_id']
    
    # Hitung statistik
=======
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    # Hapus flag setelah welcome page ditampilkan
    show_welcome = session.pop('show_welcome', False)
    
    return render_template('welcome.html', 
                         username=session.get('nama_lengkap', session.get('username', 'Pengguna')))

@app.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        user = User.query.filter_by(username=username).first()
        
        if not user:
            flash('Username atau password salah', 'error')
            return redirect(url_for('admin_login'))
        
        if user.role != 'admin':
            flash('Hanya admin yang dapat login di sini', 'error')
            return redirect(url_for('admin_login'))
        
        if user and user.check_password(password) and user.is_active:
            session['user_id'] = user.id
            session['username'] = user.username
            session['nama_lengkap'] = user.nama_lengkap or user.username
            session['role'] = user.role
            
            flash('Login admin berhasil!', 'success')
            return redirect(url_for('admin_dashboard'))
        else:
            flash('Username atau password salah', 'error')
            return redirect(url_for('admin_login'))
    
    return render_template('admin/login.html')

@app.route('/dashboard')
def dashboard():
    if 'user_id' not in session:
        flash('Silakan login terlebih dahulu', 'error')
        return redirect(url_for('login'))
    
    if session.get('role') == 'admin':
        return redirect(url_for('admin_dashboard'))
    
    user_id = session['user_id']
    periode = request.args.get('periode', 'bulan')
    
>>>>>>> 4fd9892 (final code)
    total_pesanan = Pemesanan.query.filter_by(user_id=user_id).count()
    pesanan_menunggu = Pemesanan.query.filter_by(
        user_id=user_id, 
        status='Menunggu Konfirmasi'
    ).count()
    pesanan_selesai = Pemesanan.query.filter_by(
        user_id=user_id, 
        status='Selesai'
    ).count()
<<<<<<< HEAD
    
    # Ambil pesanan terbaru
=======
>>>>>>> 4fd9892 (final code)
    pesanan_terbaru = Pemesanan.query.filter_by(user_id=user_id)\
        .order_by(Pemesanan.tanggal_pesan.desc())\
        .limit(5).all()
    
    return render_template('dashboard.html', 
                         username=session['nama_lengkap'],
                         total_pesanan=total_pesanan,
                         pesanan_menunggu=pesanan_menunggu,
                         pesanan_selesai=pesanan_selesai,
<<<<<<< HEAD
                         pesanan_terbaru=pesanan_terbaru)
=======
                         pesanan_terbaru=pesanan_terbaru,
                         is_admin=False,
                         chart_data=None,
                         status_data=None,
                         status_labels=None,
                         status_values=None,
                         periode=None)

def get_chart_data(user_id=None, periode='bulan'):
    today = datetime.today().date()
    start_date = None
    labels = []
    data_points = []
    
    if periode == 'minggu':
        start_date = today - timedelta(days=7)
        for i in range(7, -1, -1):
            date = today - timedelta(days=i)
            labels.append(date.strftime('%d/%m'))
            query = Pemesanan.query.filter(db.func.date(Pemesanan.tanggal_pesan) == date)
            if user_id:
                query = query.filter_by(user_id=user_id)
            data_points.append(query.count())
    
    elif periode == 'bulan':
        start_date = today - timedelta(days=30)
        for i in range(29, -1, -1):
            date = today - timedelta(days=i)
            labels.append(date.strftime('%d/%m'))
            query = Pemesanan.query.filter(db.func.date(Pemesanan.tanggal_pesan) == date)
            if user_id:
                query = query.filter_by(user_id=user_id)
            data_points.append(query.count())
    
    elif periode == 'kuartal':
        start_date = today - timedelta(days=90)
        for i in range(89, -1, -3):
            date = today - timedelta(days=i)
            labels.append(date.strftime('%d/%m'))
            date_range_start = date
            date_range_end = date + timedelta(days=2)
            query = Pemesanan.query.filter(
                db.func.date(Pemesanan.tanggal_pesan) >= date_range_start,
                db.func.date(Pemesanan.tanggal_pesan) <= date_range_end
            )
            if user_id:
                query = query.filter_by(user_id=user_id)
            data_points.append(query.count())
    
    elif periode == 'tahun':
        start_date = today.replace(day=1) - timedelta(days=365)
        from calendar import monthrange
        months_data = []
        for i in range(11, -1, -1):
            month_date = today.replace(day=1)
            for _ in range(i):
                if month_date.month == 1:
                    month_date = month_date.replace(year=month_date.year - 1, month=12)
                else:
                    month_date = month_date.replace(month=month_date.month - 1)
            
            month_key = month_date.strftime('%Y-%m')
            if month_key not in [m[0] for m in months_data]:
                labels.append(month_date.strftime('%b %Y'))
                query = Pemesanan.query.filter(
                    db.func.extract('year', Pemesanan.tanggal_pesan) == month_date.year,
                    db.func.extract('month', Pemesanan.tanggal_pesan) == month_date.month
                )
                if user_id:
                    query = query.filter_by(user_id=user_id)
                count = query.count()
                data_points.append(count)
                months_data.append((month_key, count))
    
    chart_data = [{'date': label, 'count': count} for label, count in zip(labels, data_points)]
    
    query_status = Pemesanan.query
    if user_id:
        query_status = query_status.filter_by(user_id=user_id)
    if start_date:
        query_status = query_status.filter(Pemesanan.tanggal_pesan >= start_date)
    
    status_counts = {
        'Menunggu Konfirmasi': query_status.filter_by(status='Menunggu Konfirmasi').count(),
        'Dikonfirmasi': query_status.filter_by(status='Dikonfirmasi').count(),
        'Diproses': query_status.filter_by(status='Diproses').count(),
        'Dikirim': query_status.filter_by(status='Dikirim').count(),
        'Selesai': query_status.filter_by(status='Selesai').count()
    }
    
    return chart_data, status_counts
>>>>>>> 4fd9892 (final code)

@app.route('/admin/dashboard')
def admin_dashboard():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
<<<<<<< HEAD
    # Statistik admin
=======
    periode = request.args.get('periode', 'bulan')
    
>>>>>>> 4fd9892 (final code)
    total_users = User.query.count()
    total_pesanan = Pemesanan.query.count()
    total_pendapatan = db.session.query(db.func.sum(Pemesanan.total_harga)).scalar() or 0
    pesanan_hari_ini = Pemesanan.query.filter(
        db.func.date(Pemesanan.tanggal_pesan) == datetime.today().date()
    ).count()
    
<<<<<<< HEAD
    today = datetime.today().strftime('%d %B %Y')
    
    return render_template('dashboard.html',
=======
    chart_data, status_counts = get_chart_data(user_id=None, periode=periode)
    
    status_labels = list(status_counts.keys())
    status_values = list(status_counts.values())
    
    today = datetime.today().strftime('%d %B %Y')
    
    return render_template('admin/dashboard.html',
>>>>>>> 4fd9892 (final code)
                         total_users=total_users,
                         total_pesanan=total_pesanan,
                         total_pendapatan=total_pendapatan,
                         pesanan_hari_ini=pesanan_hari_ini,
<<<<<<< HEAD
                         today=today)
=======
                         today=today,
                         chart_data=chart_data,
                         status_data=status_counts,
                         status_labels=status_labels,
                         status_values=status_values,
                         periode=periode)
>>>>>>> 4fd9892 (final code)

@app.route('/harga')
def harga():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
<<<<<<< HEAD
    # Ambil data galon dari database
=======
    if session.get('role') == 'admin':
        flash('Admin dapat mengelola stok di halaman Kelola Stok', 'info')
        return redirect(url_for('stok'))
    
>>>>>>> 4fd9892 (final code)
    galon_list = Galon.query.filter_by(is_available=True).all()
    
    # Jika belum ada data, gunakan default
    if not galon_list:
        galon_list = DEFAULT_GALON
        # Simpan ke database
        for g in DEFAULT_GALON:
            galon = Galon(
                jenis=g['jenis'],
                harga=g['harga'],
                stok=g['stok']
            )
            db.session.add(galon)
        db.session.commit()
        galon_list = Galon.query.all()
    
    # Tentukan apakah user adalah admin untuk menampilkan form beli galon
    is_admin = session.get('role') == 'admin'
    
    return render_template("harga.html", harga=galon_list, is_admin=is_admin)

@app.route('/pesan', methods=['POST'])
def pesan_galon():
    if 'user_id' not in session:
        flash('Silakan login terlebih dahulu', 'error')
        return redirect(url_for('login'))
    
    try:
        # Ambil data dari form
        nama = request.form['nama']
        alamat = request.form['alamat']
<<<<<<< HEAD
        jumlah = int(request.form['jumlah'])
        jenis_galon = request.form['jenis_galon']
        harga_input = int(request.form['harga'])
        catatan = request.form.get('catatan', '')
        metode_pembayaran = request.form.get('metode_pembayaran', 'Cash')
        
        # Generate kode pemesanan
        from datetime import datetime
        kode_pemesanan = f"ORD{datetime.now().strftime('%Y%m%d%H%M%S')}"
        
        # Hitung total harga
        total_harga = harga_input * jumlah
        
        # Cek stok galon
        galon = Galon.query.filter_by(jenis=jenis_galon).first()
        if galon and galon.stok < jumlah:
            flash(f'Stok {jenis_galon} tidak mencukupi. Stok tersedia: {galon.stok}', 'error')
            return redirect(url_for('harga'))
        
        # Kurangi stok jika ada
        if galon:
            galon.stok -= jumlah
            if galon.stok <= 0:
                galon.is_available = False
        
        # Simpan pemesanan ke database
        new_pesanan = Pemesanan(
            user_id=session['user_id'],
            kode_pemesanan=kode_pemesanan,
            nama_pelanggan=nama,
            alamat=alamat,
            jenis_galon=jenis_galon,
            harga_satuan=harga_input,
            jumlah=jumlah,
            total_harga=total_harga,
            catatan=catatan,
            metode_pembayaran=metode_pembayaran,
            status='Menunggu Konfirmasi'
        )
        
        # Buat record pembayaran
        pembayaran = Pembayaran(
            pemesanan=new_pesanan,
            jumlah_bayar=total_harga,
=======
        catatan = request.form.get('catatan', '')
        metode_pembayaran = request.form.get('metode_pembayaran', 'Cash')
        
        # Ambil kategori yang dipilih (multiple)
        selected_galon_ids = request.form.getlist('selected_galon')
        
        if not selected_galon_ids:
            flash('Silakan pilih minimal 1 kategori galon', 'error')
            return redirect(url_for('harga'))
        
        # Generate kode pemesanan unik dengan microsecond dan random number
        # Format: ORD + YYYYMMDDHHMMSS + millisecond (3 digit) + random (4 digit)
        max_retries = 10
        kode_pemesanan = None
        
        for attempt in range(max_retries):
            now = datetime.now()
            millisecond = now.microsecond // 1000
            random_suffix = random.randint(1000, 9999)  # 4 digit untuk lebih unik
            kode_pemesanan = f"ORD{now.strftime('%Y%m%d%H%M%S')}{millisecond:03d}{random_suffix}"
            
            # Cek apakah kode sudah ada di database
            existing = Pemesanan.query.filter_by(kode_pemesanan=kode_pemesanan).first()
            if not existing:
                break
        
        if not kode_pemesanan:
            flash('Gagal membuat kode pemesanan unik. Silakan coba lagi.', 'error')
            return redirect(url_for('harga'))
        
        total_keseluruhan = 0
        pesanan_items = []
        errors = []
        
        # Validasi dan proses setiap kategori yang dipilih
        for galon_id in selected_galon_ids:
            galon = Galon.query.get(int(galon_id))
            if not galon:
                errors.append(f'Kategori galon dengan ID {galon_id} tidak ditemukan')
                continue
            
            jumlah = int(request.form.get(f'jumlah_{galon_id}', 1))
            
            if jumlah <= 0:
                errors.append(f'Jumlah untuk {galon.jenis} harus lebih dari 0')
                continue
            
            if galon.stok < jumlah:
                errors.append(f'Stok {galon.jenis} tidak mencukupi. Stok tersedia: {galon.stok}')
                continue
            
            total_harga = galon.harga * jumlah
            total_keseluruhan += total_harga
            
            pesanan_items.append({
                'galon': galon,
                'jumlah': jumlah,
                'total_harga': total_harga
            })
        
        if errors:
            for error in errors:
                flash(error, 'error')
            return redirect(url_for('harga'))
        
        if not pesanan_items:
            flash('Tidak ada item yang valid untuk dipesan', 'error')
            return redirect(url_for('harga'))
        
        # Buat record pemesanan untuk setiap kategori dengan kode_pemesanan yang sama
        pesanan_created = []
        for item in pesanan_items:
            # Update stok
            item['galon'].stok -= item['jumlah']
            if item['galon'].stok <= 0:
                item['galon'].is_available = False
            
            # Buat record pemesanan
            new_pesanan = Pemesanan(
                user_id=session['user_id'],
                kode_pemesanan=kode_pemesanan,
                nama_pelanggan=nama,
                alamat=alamat,
                jenis_galon=item['galon'].jenis,
                harga_satuan=item['galon'].harga,
                jumlah=item['jumlah'],
                total_harga=item['total_harga'],
                catatan=catatan,
                metode_pembayaran=metode_pembayaran,
                status='Menunggu Konfirmasi'
            )
            
            db.session.add(new_pesanan)
            pesanan_created.append(new_pesanan)
        
        # Buat satu record pembayaran untuk total keseluruhan
        # Link ke pesanan pertama (bisa diubah jika perlu relasi one-to-many)
        pembayaran = Pembayaran(
            pemesanan=pesanan_created[0],
            jumlah_bayar=total_keseluruhan,
>>>>>>> 4fd9892 (final code)
            metode=metode_pembayaran,
            status='Belum Bayar'
        )
        
<<<<<<< HEAD
        db.session.add(new_pesanan)
        db.session.add(pembayaran)
        db.session.commit()
        
        flash(f'Pesanan {jenis_galon} sebanyak {jumlah} galon berhasil! Kode: {kode_pemesanan}', 'success')
=======
        db.session.add(pembayaran)
        db.session.commit()
        
        # Buat pesan sukses
        items_text = ', '.join([f"{item['galon'].jenis} ({item['jumlah']} galon)" for item in pesanan_items])
        flash(f'Pesanan berhasil! Kode: {kode_pemesanan}. Total: Rp {total_keseluruhan:,}', 'success')
>>>>>>> 4fd9892 (final code)
        return redirect(url_for('pemesanan'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
        return redirect(url_for('harga'))

@app.route('/pemesanan')
def pemesanan():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    user_id = session['user_id']
    
<<<<<<< HEAD
    # Jika admin, tampilkan semua pesanan
=======
>>>>>>> 4fd9892 (final code)
    if session.get('role') == 'admin':
        pesanan_list = Pemesanan.query.order_by(Pemesanan.tanggal_pesan.desc()).all()
    else:
        pesanan_list = Pemesanan.query.filter_by(user_id=user_id)\
            .order_by(Pemesanan.tanggal_pesan.desc()).all()
    
<<<<<<< HEAD
    return render_template("pemesanan.html", pemesanan=pesanan_list)
=======
    # Kelompokkan pesanan berdasarkan kode_pemesanan
    pesanan_grouped = {}
    for pesanan in pesanan_list:
        kode = pesanan.kode_pemesanan
        if kode not in pesanan_grouped:
            pesanan_grouped[kode] = {
                'info': pesanan,  # Info utama (tanggal, nama, alamat, dll)
                'items': [],
                'total': 0
            }
        pesanan_grouped[kode]['items'].append(pesanan)
        pesanan_grouped[kode]['total'] += pesanan.total_harga
    
    return render_template(
        "pemesanan.html",
        pemesanan=pesanan_list,
        pesanan_grouped=pesanan_grouped
    )
>>>>>>> 4fd9892 (final code)

@app.route('/beli-galon', methods=['POST'])
def beli_galon():
    if 'user_id' not in session:
        flash('Silakan login terlebih dahulu', 'error')
        return redirect(url_for('login'))
    
<<<<<<< HEAD
=======
    if session.get('role') != 'admin':
        flash('Hanya admin yang dapat menambah stok', 'error')
        return redirect(url_for('dashboard'))
    
>>>>>>> 4fd9892 (final code)
    try:
        jenis_galon = request.form['jenis_galon']
        jumlah = int(request.form['jumlah'])
        
<<<<<<< HEAD
        # Cari harga galon
        galon = Galon.query.filter_by(jenis=jenis_galon).first()
        if not galon:
            flash('Jenis galon tidak ditemukan', 'error')
            return redirect(url_for('harga'))
=======
        galon = Galon.query.filter_by(jenis=jenis_galon).first()
        if not galon:
            flash('Jenis galon tidak ditemukan', 'error')
            return redirect(url_for('stok'))
>>>>>>> 4fd9892 (final code)
        
        harga_satuan = galon.harga
        total_harga = harga_satuan * jumlah
        
<<<<<<< HEAD
        # Simpan pembelian galon
=======
>>>>>>> 4fd9892 (final code)
        pembelian = PembelianGalon(
            user_id=session['user_id'],
            jenis_galon=jenis_galon,
            jumlah=jumlah,
            harga_satuan=harga_satuan,
            total_harga=total_harga
        )
        
<<<<<<< HEAD
        # Tambah stok galon
=======
>>>>>>> 4fd9892 (final code)
        galon.stok += jumlah
        galon.is_available = True
        
        db.session.add(pembelian)
        db.session.commit()
        
<<<<<<< HEAD
        flash(f'Pembelian {jenis_galon} sebanyak {jumlah} galon berhasil ditambahkan ke stok!', 'success')
=======
        flash(f'Stok {jenis_galon} berhasil ditambahkan sebanyak {jumlah} galon!', 'success')
>>>>>>> 4fd9892 (final code)
        return redirect(url_for('stok'))
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
<<<<<<< HEAD
        return redirect(url_for('harga'))
=======
        return redirect(url_for('stok'))
>>>>>>> 4fd9892 (final code)

@app.route('/stok')
def stok():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    
    if session.get('role') != 'admin':
        flash('Hanya admin yang dapat mengakses halaman stok', 'error')
        return redirect(url_for('dashboard'))
    
    galon_list = Galon.query.all()
    pembelian_list = PembelianGalon.query.order_by(PembelianGalon.tanggal_beli.desc()).all()
    
    return render_template('stok.html', 
                         galon_list=galon_list, 
                         pembelian_list=pembelian_list)

<<<<<<< HEAD
=======
@app.route('/tambah-kategori', methods=['GET', 'POST'])
def tambah_kategori():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        try:
            jenis = request.form['jenis']
            harga = int(request.form['harga'])
            stok = int(request.form.get('stok', 0))
            deskripsi = request.form.get('deskripsi', '')
            
            existing_galon = Galon.query.filter_by(jenis=jenis).first()
            if existing_galon:
                flash(f'Kategori {jenis} sudah ada', 'error')
                return redirect(url_for('tambah_kategori'))
            
            new_galon = Galon(
                jenis=jenis,
                harga=harga,
                stok=stok,
                deskripsi=deskripsi,
                is_available=True if stok > 0 else False
            )
            
            db.session.add(new_galon)
            db.session.commit()
            
            flash(f'Kategori {jenis} berhasil ditambahkan!', 'success')
            return redirect(url_for('stok'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
            return redirect(url_for('tambah_kategori'))
    
    return render_template('admin/tambah_kategori.html')

@app.route('/edit-kategori/<int:galon_id>', methods=['GET', 'POST'])
def edit_kategori(galon_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
    galon = Galon.query.get_or_404(galon_id)
    
    if request.method == 'POST':
        try:
            jenis = request.form['jenis']
            harga = int(request.form['harga'])
            stok = int(request.form.get('stok', 0))
            deskripsi = request.form.get('deskripsi', '')
            
            if jenis != galon.jenis:
                existing_galon = Galon.query.filter_by(jenis=jenis).first()
                if existing_galon:
                    flash(f'Kategori {jenis} sudah ada', 'error')
                    return redirect(url_for('edit_kategori', galon_id=galon_id))
            
            galon.jenis = jenis
            galon.harga = harga
            galon.stok = stok
            galon.deskripsi = deskripsi
            galon.is_available = True if stok > 0 else False
            
            db.session.commit()
            
            flash(f'Kategori {jenis} berhasil diperbarui!', 'success')
            return redirect(url_for('stok'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Terjadi kesalahan: {str(e)}', 'error')
            return redirect(url_for('edit_kategori', galon_id=galon_id))
    
    return render_template('admin/edit_kategori.html', galon=galon)

>>>>>>> 4fd9892 (final code)
@app.route('/update-status/<int:pesanan_id>', methods=['POST'])
def update_status(pesanan_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        status = request.form['status']
        pesanan = Pemesanan.query.get_or_404(pesanan_id)
<<<<<<< HEAD
        
        pesanan.status = status
        
        if status == 'Selesai':
            pesanan.tanggal_selesai = datetime.utcnow()
        elif status == 'Dikonfirmasi':
            pesanan.tanggal_konfirmasi = datetime.utcnow()
        
        db.session.commit()
        flash(f'Status pesanan {pesanan.kode_pemesanan} berhasil diupdate menjadi {status}', 'success')
=======
        kode_pemesanan = pesanan.kode_pemesanan
        
        # Update status untuk semua item dengan kode_pemesanan yang sama
        semua_pesanan = Pemesanan.query.filter_by(kode_pemesanan=kode_pemesanan).all()
        
        for p in semua_pesanan:
            p.status = status
            
            if status == 'Selesai':
                p.tanggal_selesai = datetime.utcnow()
            elif status == 'Dikonfirmasi':
                p.tanggal_konfirmasi = datetime.utcnow()
        
        db.session.commit()
        flash(f'Status pesanan {kode_pemesanan} berhasil diupdate menjadi {status}', 'success')
>>>>>>> 4fd9892 (final code)
        
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return redirect(url_for('pemesanan'))
<<<<<<< HEAD
=======

@app.route('/delete/pemesanan/<int:pesanan_id>', methods=['POST'])
def delete_pemesanan(pesanan_id):
    if 'user_id' not in session:
        flash('Silakan login terlebih dahulu', 'error')
        return redirect(url_for('login'))
    
    pesanan = Pemesanan.query.get_or_404(pesanan_id)
    
    if session.get('role') != 'admin' and pesanan.user_id != session['user_id']:
        flash('Akses ditolak', 'error')
        return redirect(url_for('pemesanan'))
    
    try:
        kode_pemesanan = pesanan.kode_pemesanan
        
        # Hapus semua item dengan kode_pemesanan yang sama
        semua_pesanan = Pemesanan.query.filter_by(kode_pemesanan=kode_pemesanan).all()
        
        for p in semua_pesanan:
            # Kembalikan stok
            galon = Galon.query.filter_by(jenis=p.jenis_galon).first()
            if galon:
                galon.stok += p.jumlah
                galon.is_available = True
            
            # Hapus pembayaran terkait (jika ada)
            Pembayaran.query.filter_by(pemesanan_id=p.id).delete()
            
            # Hapus pesanan
            db.session.delete(p)
        
        db.session.commit()
        flash(f'Pesanan {kode_pemesanan} berhasil dihapus', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return redirect(url_for('pemesanan'))

@app.route('/delete/galon/<int:galon_id>', methods=['POST'])
def delete_galon(galon_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('stok'))
    
    try:
        galon = Galon.query.get_or_404(galon_id)
        
        if Pemesanan.query.filter_by(jenis_galon=galon.jenis).count() > 0:
            flash('Tidak bisa menghapus galon yang sudah pernah dipesan', 'error')
            return redirect(url_for('stok'))
        
        PembelianGalon.query.filter_by(jenis_galon=galon.jenis).delete()
        db.session.delete(galon)
        db.session.commit()
        
        flash(f'Galon {galon.jenis} berhasil dihapus', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return redirect(url_for('stok'))

@app.route('/delete/user/<int:user_id>', methods=['POST'])
def delete_user(user_id):
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
    if user_id == session['user_id']:
        flash('Tidak bisa menghapus akun sendiri', 'error')
        return redirect(url_for('admin_dashboard'))
    
    try:
        user = User.query.get_or_404(user_id)
        if user.role == 'admin':
            flash('Tidak bisa menghapus admin lain', 'error')
            return redirect(url_for('admin_dashboard'))
        
        Pemesanan.query.filter_by(user_id=user_id).update({'user_id': None})
        PembelianGalon.query.filter_by(user_id=user_id).delete()
        db.session.delete(user)
        db.session.commit()
        
        flash(f'User {user.username} berhasil dihapus', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Terjadi kesalahan: {str(e)}', 'error')
    
    return redirect(url_for('admin_dashboard'))

@app.route('/report/excel')
def report_excel():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
    if not OPENPYXL_AVAILABLE:
        flash('Fitur Excel tidak tersedia. Silakan install openpyxl: pip install openpyxl', 'error')
        return redirect(url_for('dashboard'))
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Pemesanan"
    
    # Header/Kop Laporan
    title_fill = PatternFill(start_color="1e40af", end_color="3b82f6", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=16)
    subtitle_font = Font(bold=True, size=12)
    info_font = Font(size=10)
    
    # Baris 1: Nama Perusahaan
    ws.merge_cells('A1:I1')
    cell_title = ws.cell(row=1, column=1)
    cell_title.value = "ARTA TIRTA - DEPOT AIR MINUM ISI ULANG"
    cell_title.fill = title_fill
    cell_title.font = title_font
    cell_title.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30
    
    # Baris 2: Judul Laporan
    ws.merge_cells('A2:I2')
    cell_subtitle = ws.cell(row=2, column=1)
    cell_subtitle.value = "LAPORAN PEMESANAN GALON"
    cell_subtitle.font = subtitle_font
    cell_subtitle.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25
    
    # Baris 3: Informasi Tanggal
    ws.merge_cells('A3:I3')
    cell_date = ws.cell(row=3, column=1)
    cell_date.value = f"Dicetak pada: {datetime.now().strftime('%d %B %Y %H:%M:%S')}"
    cell_date.font = info_font
    cell_date.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[3].height = 20
    
    # Baris 4: Kosong untuk spacing
    ws.row_dimensions[4].height = 10
    
    # Baris 5: Header Tabel
    header_fill = PatternFill(start_color="1e40af", end_color="3b82f6", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    headers = ['Kode Pesanan', 'Tanggal', 'Nama Pelanggan', 'Jenis Galon', 'Jumlah', 'Harga Satuan', 'Total Harga', 'Status', 'Metode Pembayaran']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[5].height = 25
    
    pesanan_list = Pemesanan.query.order_by(Pemesanan.tanggal_pesan.desc()).all()
    
    # Hitung total pendapatan
    total_pendapatan = sum(p.total_harga for p in pesanan_list)
    
    for row_num, pesanan in enumerate(pesanan_list, 6):
        ws.cell(row=row_num, column=1, value=pesanan.kode_pemesanan)
        ws.cell(row=row_num, column=2, value=pesanan.tanggal_pesan.strftime('%d-%m-%Y %H:%M'))
        ws.cell(row=row_num, column=3, value=pesanan.nama_pelanggan)
        ws.cell(row=row_num, column=4, value=pesanan.jenis_galon)
        ws.cell(row=row_num, column=5, value=pesanan.jumlah)
        ws.cell(row=row_num, column=6, value=pesanan.harga_satuan)
        ws.cell(row=row_num, column=7, value=pesanan.total_harga)
        ws.cell(row=row_num, column=8, value=pesanan.status)
        ws.cell(row=row_num, column=9, value=pesanan.metode_pembayaran)
    
    # Baris terakhir: Total Pendapatan
    last_row = 5 + len(pesanan_list) + 1
    ws.merge_cells(f'A{last_row}:F{last_row}')
    cell_total_label = ws.cell(row=last_row, column=1)
    cell_total_label.value = "TOTAL PENDAPATAN:"
    cell_total_label.font = Font(bold=True, size=12)
    cell_total_label.alignment = Alignment(horizontal='right', vertical='center')
    
    ws.merge_cells(f'G{last_row}:I{last_row}')
    cell_total = ws.cell(row=last_row, column=7)
    cell_total.value = f"Rp {total_pendapatan:,}"
    cell_total.font = Font(bold=True, size=12, color="1e40af")
    cell_total.alignment = Alignment(horizontal='center', vertical='center')
    cell_total.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
    ws.row_dimensions[last_row].height = 25
    
    # Auto-adjust column width
    from openpyxl.utils import get_column_letter
    from openpyxl.cell.cell import MergedCell
    
    # Loop melalui semua kolom berdasarkan index
    for col_idx in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        # Loop melalui semua baris di kolom ini (mulai dari baris 5 karena header di baris 5)
        for row_idx in range(5, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            # Skip MergedCell objects
            if not isinstance(cell, MergedCell):
                try:
                    if cell.value is not None:
                        cell_value = str(cell.value)
                        if len(cell_value) > max_length:
                            max_length = len(cell_value)
                except:
                    pass
        
        # Set width
        if max_length > 0:
            adjusted_width = min(max_length + 2, 50)
        else:
            adjusted_width = 12  # Default width
        ws.column_dimensions[column_letter].width = adjusted_width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"laporan_pemesanan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)

@app.route('/report/pdf')
def report_pdf():
    if 'user_id' not in session or session.get('role') != 'admin':
        flash('Akses ditolak. Hanya untuk admin.', 'error')
        return redirect(url_for('dashboard'))
    
    if not REPORTLAB_AVAILABLE:
        flash('Fitur PDF tidak tersedia. Silakan install reportlab: pip install reportlab', 'error')
        return redirect(url_for('dashboard'))
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []
    
    styles = getSampleStyleSheet()
    
    # Header/Kop Laporan
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    
    # Style untuk header
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Heading1'],
        fontSize=20,
        textColor=colors.HexColor('#1e40af'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#3b82f6'),
        spaceAfter=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    info_style = ParagraphStyle(
        'CustomInfo',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#64748b'),
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    # Nama Perusahaan
    title = Paragraph("ARTA TIRTA", header_style)
    elements.append(title)
    
    # Tagline
    tagline = Paragraph("DEPOT AIR MINUM ISI ULANG", subtitle_style)
    elements.append(tagline)
    
    # Judul Laporan
    report_title = Paragraph("LAPORAN PEMESANAN GALON", subtitle_style)
    elements.append(report_title)
    elements.append(Spacer(1, 0.1*inch))
    
    # Informasi Tanggal
    date_info = Paragraph(f"Dicetak pada: {datetime.now().strftime('%d %B %Y %H:%M:%S')}", info_style)
    elements.append(date_info)
    elements.append(Spacer(1, 0.2*inch))
    
    data = [['Kode', 'Tanggal', 'Pelanggan', 'Jenis Galon', 'Jml', 'Harga', 'Total', 'Status']]
    
    pesanan_list = Pemesanan.query.order_by(Pemesanan.tanggal_pesan.desc()).all()
    for pesanan in pesanan_list:
        data.append([
            pesanan.kode_pemesanan,
            pesanan.tanggal_pesan.strftime('%d-%m-%Y'),
            pesanan.nama_pelanggan[:20],
            pesanan.jenis_galon,
            str(pesanan.jumlah),
            f"Rp {pesanan.harga_satuan:,}",
            f"Rp {pesanan.total_harga:,}",
            pesanan.status
        ])
    
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e40af')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Hitung total pendapatan
    total_pendapatan = sum(p.total_harga for p in pesanan_list)
    
    elements.append(Spacer(1, 0.2*inch))
    
    # Total Pendapatan dengan styling menggunakan Table
    total_style = ParagraphStyle(
        'TotalStyle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#1e40af'),
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Buat tabel untuk total dengan background
    total_data = [[f"TOTAL PENDAPATAN: Rp {total_pendapatan:,}"]]
    total_table = Table(total_data, colWidths=[7*inch])
    total_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#dbeafe')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e40af')),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('TOPPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 2, colors.HexColor('#1e40af')),
    ]))
    elements.append(total_table)
    
    doc.build(elements)
    buffer.seek(0)
    
    filename = f"laporan_pemesanan_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)
>>>>>>> 4fd9892 (final code)

@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout', 'info')
    return redirect(url_for('login'))

<<<<<<< HEAD
# ⭐⭐⭐ FUNGSI BARU UNTUK MIGRASI DATA ⭐⭐⭐
def backup_sqlite_to_json():
    """Backup data dari SQLite ke file JSON"""
=======
def backup_sqlite_to_json():
>>>>>>> 4fd9892 (final code)
    import sqlite3
    import json
    
    if not os.path.exists('depot_air.db'):
        print("File depot_air.db tidak ditemukan, skip backup")
        return
    
    conn = sqlite3.connect('depot_air.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    
<<<<<<< HEAD
    # Ambil semua tabel
=======
>>>>>>> 4fd9892 (final code)
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
    tables = cursor.fetchall()
    
    data = {}
    for table in tables:
        table_name = table[0]
        cursor.execute(f"SELECT * FROM {table_name}")
        rows = cursor.fetchall()
        data[table_name] = [dict(row) for row in rows]
    
<<<<<<< HEAD
    # Simpan ke JSON
=======
>>>>>>> 4fd9892 (final code)
    with open('backup_data.json', 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, default=str)
    
    conn.close()
<<<<<<< HEAD
    print("✅ Backup SQLite berhasil disimpan ke backup_data.json")

def migrate_to_mysql():
    """Pindahkan data dari SQLite ke MySQL"""
    backup_sqlite_to_json()
    
    if not os.path.exists('backup_data.json'):
        print("❌ File backup_data.json tidak ditemukan")
        return
    
    try:
        # Baca data backup
        with open('backup_data.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print("📦 Mulai migrasi data ke MySQL...")
        
        # Urutan migrasi
=======
    print("Backup SQLite berhasil disimpan ke backup_data.json")

def migrate_to_mysql():
    backup_sqlite_to_json()
    
    if not os.path.exists('backup_data.json'):
        print("File backup_data.json tidak ditemukan")
        return
    
    try:
        with open('backup_data.json', 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        print("Mulai migrasi data ke MySQL...")
        
>>>>>>> 4fd9892 (final code)
        order = ['user', 'galon', 'pemesanan', 'pembelian_galon', 'pembayaran']
        
        for table_name in order:
            if table_name not in data or not data[table_name]:
                continue
            
            print(f"  Migrasi tabel: {table_name}")
            
<<<<<<< HEAD
            # Get model class
=======
>>>>>>> 4fd9892 (final code)
            model_class = None
            if table_name == 'user':
                model_class = User
            elif table_name == 'galon':
                model_class = Galon
            elif table_name == 'pemesanan':
                model_class = Pemesanan
            elif table_name == 'pembelian_galon':
                model_class = PembelianGalon
            elif table_name == 'pembayaran':
                model_class = Pembayaran
            
            if not model_class:
                continue
            
<<<<<<< HEAD
            # Insert data
            for row_data in data[table_name]:
                try:
                    # Hapus id untuk auto increment
                    if 'id' in row_data:
                        del row_data['id']
                    
                    # Buat objek
=======
            for row_data in data[table_name]:
                try:
                    if 'id' in row_data:
                        del row_data['id']
                    
>>>>>>> 4fd9892 (final code)
                    obj = model_class(**row_data)
                    db.session.add(obj)
                    
                except Exception as e:
                    print(f"    Error: {e}")
                    continue
            
            db.session.commit()
<<<<<<< HEAD
            print(f"  ✅ {len(data[table_name])} data berhasil dimigrasi")
        
        print("🎉 Migrasi data SELESAI!")
        
    except Exception as e:
        db.session.rollback()
        print(f"❌ Error migrasi: {e}")

def init_database():
    """Inisialisasi database dengan data default"""
    with app.app_context():
        # Buat semua tabel di MySQL
        db.create_all()
        print("✅ Tabel berhasil dibuat di MySQL")
        
        # Cek apakah sudah ada data
=======
            print(f"  {len(data[table_name])} data berhasil dimigrasi")
        
        print("Migrasi data SELESAI!")
        
    except Exception as e:
        db.session.rollback()
        print(f"Error migrasi: {e}")

def remove_unique_constraint_if_exists():
    """Menghapus constraint UNIQUE dari kode_pemesanan jika ada"""
    try:
        with app.app_context():
            # Cek apakah ada constraint unique pada kode_pemesanan
            from sqlalchemy import inspect, text
            inspector = inspect(db.engine)
            
            # Cek constraint di tabel pemesanan
            try:
                # Coba hapus index/constraint jika ada
                with db.engine.connect() as conn:
                    # Cek index yang ada
                    result = conn.execute(text("SHOW INDEX FROM pemesanan WHERE Column_name = 'kode_pemesanan' AND Non_unique = 0"))
                    indexes = result.fetchall()
                    
                    for index in indexes:
                        index_name = index[2]  # Key_name
                        if index_name != 'PRIMARY':
                            try:
                                conn.execute(text(f"ALTER TABLE pemesanan DROP INDEX {index_name}"))
                                conn.commit()
                                print(f"Constraint UNIQUE '{index_name}' pada kode_pemesanan berhasil dihapus")
                            except Exception as e:
                                print(f"Tidak dapat menghapus constraint: {e}")
            except Exception as e:
                print(f"Error saat cek constraint: {e}")
    except Exception as e:
        print(f"Error saat menghapus constraint: {e}")

def update_kode_pemesanan_column_length():
    """Memperpanjang kolom kode_pemesanan dari 20 menjadi 30 karakter"""
    try:
        with app.app_context():
            from sqlalchemy import text
            with db.engine.connect() as conn:
                try:
                    # Cek panjang kolom saat ini
                    result = conn.execute(text("""
                        SELECT CHARACTER_MAXIMUM_LENGTH 
                        FROM INFORMATION_SCHEMA.COLUMNS 
                        WHERE TABLE_SCHEMA = DATABASE() 
                        AND TABLE_NAME = 'pemesanan' 
                        AND COLUMN_NAME = 'kode_pemesanan'
                    """))
                    current_length = result.fetchone()
                    
                    if current_length and current_length[0] < 30:
                        # Ubah panjang kolom menjadi 30
                        conn.execute(text("ALTER TABLE pemesanan MODIFY COLUMN kode_pemesanan VARCHAR(30) NOT NULL"))
                        conn.commit()
                        print(f"Kolom kode_pemesanan berhasil diperpanjang menjadi 30 karakter")
                    else:
                        print(f"Kolom kode_pemesanan sudah memiliki panjang yang cukup ({current_length[0] if current_length else 'N/A'})")
                except Exception as e:
                    # Jika tabel belum ada, akan dibuat dengan panjang yang benar
                    print(f"Tabel pemesanan mungkin belum ada atau error: {e}")
    except Exception as e:
        print(f"Error saat update kolom: {e}")

def init_database():
    with app.app_context():
        # Hapus constraint unique jika ada (karena satu pesanan bisa punya multiple items)
        remove_unique_constraint_if_exists()
        
        # Update panjang kolom kode_pemesanan jika perlu
        update_kode_pemesanan_column_length()
        
        db.create_all()
        print("Tabel berhasil dibuat di MySQL")
        
>>>>>>> 4fd9892 (final code)
        user_count = User.query.count()
        galon_count = Galon.query.count()
        
        if user_count == 0 and galon_count == 0:
<<<<<<< HEAD
            print("📦 Migrasi data dari SQLite (jika ada)...")
            migrate_to_mysql()
        
        # Cek apakah admin sudah ada
=======
            print("Migrasi data dari SQLite (jika ada)...")
            migrate_to_mysql()
        
>>>>>>> 4fd9892 (final code)
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin_user = User(
                username='admin',
                nama_lengkap='Administrator',
                email='admin@artatirta.com',
                role='admin'
            )
            admin_user.set_password('12345')
            db.session.add(admin_user)
<<<<<<< HEAD
            print("✅ Admin user dibuat")
        
        # Cek apakah galon sudah ada
=======
            print("Admin user dibuat")
        
>>>>>>> 4fd9892 (final code)
        galon = Galon.query.first()
        if not galon:
            for g in DEFAULT_GALON:
                galon_item = Galon(
                    jenis=g['jenis'],
                    harga=g['harga'],
                    stok=g['stok']
                )
                db.session.add(galon_item)
<<<<<<< HEAD
            print("✅ Data galon default dibuat")
        
        try:
            db.session.commit()
            print('🎉 Database siap digunakan!')
        except Exception as e:
            db.session.rollback()
            print(f'❌ Error: {e}')

if __name__ == '__main__':
    # ⭐⭐⭐ INI YANG PENTING ⭐⭐⭐
    # Koneksi ke database
    db.init_app(app)
    
    # Inisialisasi database
    print("🔧 Memulai inisialisasi database...")
    init_database()
    
    # Jalankan aplikasi
    print("🚀 Aplikasi berjalan di http://localhost:5000")
=======
            print("Data galon default dibuat")
        
        try:
            db.session.commit()
            print('Database siap digunakan!')
        except Exception as e:
            db.session.rollback()
            print(f'Error: {e}')

if __name__ == '__main__':
    db.init_app(app)
    print("Memulai inisialisasi database...")
    init_database()
    print("Aplikasi berjalan di http://localhost:5000")
>>>>>>> 4fd9892 (final code)
    app.run(debug=True, port=5000)